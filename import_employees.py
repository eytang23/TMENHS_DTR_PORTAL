import sqlite3
import pandas as pd
from openpyxl import load_workbook

EXCEL_FILE = "Punch Time(20260601-20260630).xlsx"     # pangalan ng excel mo

# ==========================
# LOAD EXCEL (same sa DTR generator)
# ==========================

wb = load_workbook(EXCEL_FILE)
ws = wb.active

data = []

for i, row in enumerate(ws.iter_rows(values_only=True)):

    if i < 2:
        continue

    if row[0] is None:
        continue

    data.append(row)

headers = [
    str(cell.value).strip() if cell.value else ""
    for cell in ws[2]
]

df = pd.DataFrame(data, columns=headers)

print(df.columns.tolist())
print(df.head())

# ==========================
# CONNECT DATABASE
# ==========================

conn = sqlite3.connect("database.db")
cursor = conn.cursor()

for _, row in df.iterrows():

    employee_id = str(row["Person ID"]).strip()
    name = str(row["Person Name"]).strip()

    cursor.execute("""
        INSERT OR IGNORE INTO employees
        (employee_id, name, passcode)
        VALUES (?,?,?)
    """, (
        employee_id,
        name,
        employee_id
    ))

conn.commit()
conn.close()

print("Employees Imported Successfully!")